##The standard swerlein algorithm, as he wrote it. has the basic corrections
##and verified against the testpoint code
##02/12/2016

import visa
import csv
import numpy as np
import time as time
import matplotlib.pyplot as plt
import gui
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()
ws = wb.active

def FNFreq(self,Expect):
    self.instrument.write("TARM HOLD;LFILTER ON;LEVEL 0,DC;FSOURCE ACDCV")
    if self.CodeValidation.QueryValidate.GetValue()==True:
        tarm = str(self.instrument.query('TARM?')) #Check to see if 3458A is set to 'HOLD' mode
        query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store TARM? validation in txt file.
        query_file.write("\nQuery : TARM? = " + tarm)
        if tarm[0]==str(4):
            query_file.write("---> 3458A Triggering Is Disabled. Matches With Python. \n")
        else:
            query_file.write("---> 3458A Triggering Is Not Disabled. Doesnt Match With Python. \n")
        query_file.close()
        lfilter = str(self.instrument.query('LFILTER?')) #Check to see if 3458A is set to 'LFILTER' mode
        query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store LFILTER? validation in txt file.
        query_file.write("\nQuery : LFILTER? = " + lfilter)
        if lfilter[0]==str(1):
            query_file.write("---> 3458A LFILTER? Matches With Python \n")
        else:
            query_file.write("---> 3458A LFILTER? Does Not Match Python \n")
        query_file.close()
        level = str(self.instrument.query('LEVEL?')) #Check to see if 3458A is set to 'LEVEL' mode
        query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store LEVEL? validation in txt file.
        query_file.write("\nQuery : LEVEL? = " + level)
        if str(1) in level:
            query_file.write("---> 3458A LEVEL? Set To DC Mode. Matches With Python. \n")
        else:
            query_file.write("---> 3458A LEVEL? Not Set To DC Mode. Doesnt Match With Python. \n")
        query_file.close()
        fsource = str(self.instrument.query('FSOURCE?')) #Check to see if 3458A is set to 'FSOURCE' mode
        query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store FSOURCE? validation in txt file.
        query_file.write("\nQuery : FSOURCE? = " + fsource)
        if fsource[0]==str(3):
            query_file.write("---> 3458A FSOURCE? Set to ACDCV Mode. Matches With Python. \n")
        else:
            query_file.write("---> 3458A FSOURCE? Wasnt Set to ACDCV Mode. Doesnt Match With Python. \n")
        query_file.close()
    self.instrument.write("FREQ "+str(Expect*0.9))
    Cal = float(self.instrument.query("CAL? 245"))
    Freq=Expect/Cal
    return Freq

def FNVmeter_bw(Freq,Range):
    Lvfilter = 120000.0  #LOW VOLTAGE INPUT FILTER B.W.
    Hvattn=36000.0     #HIGH VOLTAGE ATTENUATOR B.W.(NUMERATOR)
    Gain100bw=82000.0   #AMP GAIN 100 B.W. PEAKING CORRECTION!
    if Range<=0.12:
        Bw_corr=(1+(Freq/Lvfilter)**2)/(1+(Freq/Gain100bw)**2)
        Bw_corr=np.sqrt(Bw_corr)
    elif Range<=12:
        Bw_corr=(1+(Freq/Lvfilter)**2)
        Bw_corr=np.sqrt(Bw_corr)
    elif Range>12:
        Bw_corr=(1+(Freq/Hvattn)**2)
        Bw_corr=np.sqrt(Bw_corr)
        
    return Bw_corr

def OnStart (self,port):
    rm = visa.ResourceManager()
    self.instrument = rm.open_resource('GPIB0::'+str(port)+'::INSTR')
    self.instrument.timeout = 10000
    self.readings = []
    self.times = []
    self.AcdcArray = []
    self.AcArray = []
    self.MeanArray = []
    self.MemArray = []

## def offsettimes(times):
##     array = np.array(times)
##     initial = times[0]
##     return array-initial

## def save(self,name,Acdc,Ac,Mean,Mem):
##     if name=="":
##         values = time.localtime(time.time())
##         file_name = str(values[0])+'.'+str(values[1])+'.'+str(values[2])+'.'+str(values[3])+'.'+str(values[4])+".csv"
##     else:
##         file_name = name
##     tosave = []
##     tosave.append(offsettimes(self.times))
##     if Acdc == True:
##         tosave.append(self.AcdcArray)
##     if Ac == True:
##         tosave.append(self.AcArray)
##     if Mean == True:
##         tosave.append(self.MeanArray)
##     tosave = np.array(tosave)
##     np.savetxt(file_name,tosave.transpose())

## def SaveMem(data):
##     values = time.localtime(time.time())
##     file_name = 'data.'+str(values[0])+'.'+str(values[1])+'.'+str(values[2])+'.'+str(values[3])+'.'+str(values[4])+".csv"
##     data = np.array(data)
##     np.savetxt(file_name,data)
##     #data.tofile(file_name,sep = ',')
##     
## def ReturnData(self):
##     return [offsettimes(self.times),self.readings]


## def SavePlot(self,name):
##     if name=="":
##         values = time.localtime(time.time())
##         file_name = str(values[0])+'.'+str(values[1])+'.'+str(values[2])+'.'+str(values[3])+'.'+str(values[4])
##     else:
##         file_name = name
##     plt.plot(offsettimes(self.times),self.readings)
##     plt.ylabel('Voltage [V]')
##     plt.xlabel('Time [ms]')
##     plt.savefig(str(file_name)+'.pdf')
    

def reset(self):
    self.instrument.timeout = 30000
    self.instrument.write('DISP OFF, RESET')
    self.instrument.write('RESET')
    self.instrument.write('end 2')
    if self.CodeValidation.QueryValidate.GetValue()==True:
        end = str(self.instrument.query('END?')) #Check to see if 3458A is set to 'END' mode
        query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store END? validation in txt file.
        query_file.write("\nQuery : END? = " + end)
        if end[0]==str(2):
            query_file.write("---> 3458A Ready For Next Measurement \n")
        else:
            query_file.write("---> 3458A Not Ready For Next Measurement \n")
        query_file.close()
    self.instrument.write('DISP OFF, READY')
    
def run(self,Nharm_set,Nbursts_set,Acdc,Ac,MeanV,row_inc,filename):
    reset(self)

    Force = self.ForceParameters.GetValue() #input("force True/False: ")             #are variables forced or not
    Meas_time=float(self.MeasureTime.GetValue())              # Target measure time
    Tsampforce=0.001           #  FORCED PARAMETER
    Aperforce=Tsampforce-(3e-5)#  FORCED PARAMETER
    Numforce=800.0              #  FORCED PARAMETER
    Aper_targ=0.001            #A/D APERTURE TARGET (SEC)
    Nharm_min=Nharm_set                   #MINIMUM # HARMONICS SAMPLED BEFORE ALIAS
    Nbursts=Nbursts_set                 #NUMBER OF BURSTS USED FOR EACH MEASUREMENT



    if (Force==False):
        #determine input signal RMS,range and frequency
        RMS = float(self.instrument.query('ACV')) #read value of AC voltage
        Range = 1.55*RMS
        self.instrument.write('DISP OFF, FREQUENCY')
        Expect_Freq = float(self.instrument.query('FREQ')) #read frequency of signal
        Freq = FNFreq(self,Expect_Freq)
        #SAMP PARAM
        Aper=Aper_targ
        Tsamp=(1e-7)*int((Aper+(3e-5))/(1e-7)+0.5) #rounds to 100ns
        #print("first Tsamp: "+str(Tsamp))
        Submeas_time=Meas_time/Nbursts #meas_time specified, this is target time per burst
        Burst_time=Submeas_time*Tsamp/(0.0015+Tsamp) #IT TAKES 1.5ms FOR EACH sample to compute Sdev
        Approxnum=int(Burst_time/Tsamp+0.5)
        #print(Approxnum)
        Ncycle=int(Burst_time*Freq+0.5) # NUMBER OF 1/Freq TO SAMPLE
        #print(" ")
        #print("Ncycle: "+str(Ncycle))
        if Ncycle==0:
            print("Ncycle was 0, set to 1")
            Ncycle=1
            Tsamp=(1e-7)*int(1.0/Freq/Approxnum/(1e-7)+0.5)
            Nharm=int(1.0/Tsamp/2.0/Freq)
            #print("Nharm: "+str(Nharm))
            if Nharm<Nharm_min:
                #print("Nharm too small, set to 6")
                Nharm=Nharm_min
                Tsamp=(1e-7)*int(1.0/2.0/Nharm/Freq/(1e-7)+0.5)
        else:
            Nharm=int(1/Tsamp/2/Freq)
            #print("Nharm: "+str(Nharm))
            if Nharm<Nharm_min:
                Nharm = Nharm_min
                #print("Nharm too small, set to 6")

            Tsamptemp=(1e-7)*int(1.0/2.0/Nharm/Freq/(1e-7)+0.5)
            Burst_time=Submeas_time*Tsamptemp/(0.0015+Tsamptemp)
            Ncycle=int(Burst_time*Freq+1) ##0.5 to 1

            Num=int(Ncycle/Freq/Tsamptemp+0.5)
            #print("Num: "+str(Num))
            
            if Ncycle>1:
                K=int(Num/20/Nharm+1) #0.5 to 1
                #print("K= "+str(K))
            else:
                K=0
                #print("K=0")
            Tsamp=(1e-7)*int(Ncycle/Freq/(Num-K)/(1e-7)+0.5)
            if Tsamp-Tsamptemp<(1e-7):
                #print("Tsamp increased from "+str(Tsamp)+"to "+str((Tsamp+1e-7)))
                Tsamp=Tsamp+1e-7
                
        Aper=Tsamp-(3e-5)
        Num=int(Ncycle/Freq/Tsamp+0.5)
        #print('NEW NUM '+str(Num))
        if Aper>1.0:
            Aper = 1.0
            print("Aperture too large, automatically set to 1")
        elif Aper<1e-6:
            print("A/D APERTURE IS TOO SMALL")
            print("LOWER Aper_targ, Nharm, OR INPUT Freq")
            print("Aperture set to 1e-6")
            Aper = 1e-6
    else:
        print("Using Forced Parameters")
        #Freq = input("Input signal frequency: ")
        Freq=float(self.ForceFreq.GetValue()) #Forced frequency for testing.
        RMS = float(self.instrument.query('ACV')) #read value of AC voltage
        if self.CodeValidation.QueryValidate.GetValue()==True:
            func = str(self.instrument.query('FUNC?')) #Check to see if 3458A is set to 'ACV' mode
            query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store FUNC? validation in txt file.
            query_file.write("\nQuery : FUNC? = " + func)
            if func[0]==str(2):
                query_file.write("---> 3458A Set To ACV Mode. Matches With Python. \n")
            else:
                query_file.write("---> 3458A Not Set To ACV Mode. Doesnt Match With Python. \n")
            query_file.close()
        Range=1.55*RMS
        #Range = 1.55*float(input("RMS value of AC signal: "))
        Tsamp=Tsampforce
        Aper=Aperforce
        Num=Numforce

    ######   setup machine
    self.instrument.write('DISP OFF,SETUP')
    self.instrument.write('TARM HOLD;AZERO OFF;DCV '+str(Range))
    if self.CodeValidation.QueryValidate.GetValue()==True:
        tarm = str(self.instrument.query('TARM?')) #Check to see if 3458A is set to 'HOLD' mode
        query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store TARM? validation in txt file.
        query_file.write("\nQuery : TARM? = " + tarm)
        if tarm[0]==str(4):
            query_file.write("---> 3458A Triggering Is Disabled. Matches With Python. \n")
        else:
            query_file.write("---> 3458A Triggering Is Not Disabled. Doesnt Match With Python. \n")
        query_file.close()
        azero = str(self.instrument.query('AZERO?')) #Check to see if 3458A is set to 'AZERO' mode
        query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store AZERO? validation in txt file.
        query_file.write("\nQuery : AZERO? = " + azero)
        if azero[0]==str(0):
            query_file.write("---> 3458A AZERO? only once per func/range/aper/NPLC/res change. Matches With Python. \n")
        else:
            query_file.write("---> 3458A NOT AZERO? only once per func/range/aper/NPLC/res change. Doesnt Match With Python. \n")
        query_file.close()
        func = str(self.instrument.query('FUNC?')) #Check to see if 3458A is set to 'DCV' mode
        query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store FUNC? validation in txt file.
        query_file.write("\nQuery : FUNC? = " + func)
        if func[0]==str(1):
            query_file.write("---> 3458A Set To DCV Mode With Range And RMS="+str(RMS)+". Matches With Python.\n")
        else:
            query_file.write("---> 3458A Not Set To DCV Mode With Range And RMS="+str(RMS)+". Doesnt Match With Python.\n")
        query_file.close()
    self.instrument.write('APER '+str(Aper))
    if self.CodeValidation.QueryValidate.GetValue()==True:
        aper = float(self.instrument.query('APER?')) #Check to see if 3458A is set to 'APER' mode
        query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store APER? validation in txt file.
        query_file.write("\nQuery : APER? = " + str(aper) + " \n")
        query_file.write("---> 3458A Compare if APER? matches with Python="+str(Aper) +"\n")
        query_file.close()
    self.instrument.write('TIMER '+str(Tsamp))
    if self.CodeValidation.QueryValidate.GetValue()==True:
        timer = float(self.instrument.query('TIMER?')) #Check to see if 3458A is set to 'TIMER' mode
        query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store TIMER? validation in txt file.
        query_file.write("\nQuery : TIMER? = " + str(timer) + " \n")
        query_file.write("---> 3458A Compare if TIMER? matches with Python="+str(Tsamp) +"\n")
        query_file.close()
    self.instrument.write('NRDGS '+str(Num)+',TIMER')
    if self.CodeValidation.QueryValidate.GetValue()==True:
        nrdgs = str(self.instrument.query('NRDGS?')) #Check to see if 3458A is set to 'NRDGS' mode
        query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store NRDGS? validation in txt file.
        query_file.write("\nQuery : NRDGS? = " + nrdgs)
        if str(int(Num)) in nrdgs:
            query_file.write("---> 3458A NRDGS? Matches With Python="+str(Num) +"\n")
        else:
            query_file.write("---> 3458A NRDGS? Doesnt Match With Python. \n")
        query_file.close()
    #print("TSAMP IS "+str(Tsamp))
    #self.instrument.write('SWEEP '+str(Tsamp)+','+str(Num))
    self.instrument.write('TRIG LEVEL;LEVEL 0,DC;DELAY 0;LFILTER ON')
    if self.CodeValidation.QueryValidate.GetValue()==True:
        trig = str(self.instrument.query('TRIG?')) #Check to see if 3458A is set to 'TRIG' mode
        query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store TRIG? validation in txt file.
        query_file.write("\nQuery : TRIG? = " + trig)
        if trig[0]==str(7):
            query_file.write("---> 3458A Set To TRIG? LEVEL Mode. Matches With Python. \n")
        else:
            query_file.write("---> 3458A Not Set To TRIG? LEVEL Mode. Doesnt Match With Python. \n")
        query_file.close()
        level = str(self.instrument.query('LEVEL?')) #Check to see if 3458A is set to 'LEVEL' mode
        query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store LEVEL? validation in txt file.
        query_file.write("\nQuery : LEVEL? = " + level)
        if str(1) in level:
            query_file.write("---> 3458A LEVEL? Set To DC Mode. Matches With Python. \n")
        else:
            query_file.write("---> 3458A LEVEL? Not To DC Mode. Doesnt Match With Python. \n")
        query_file.close()
        delay = str(self.instrument.query('DELAY?')) #Check to see if 3458A is set to 'DELAY' mode
        query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store DELAY? validation in txt file.
        query_file.write("\nQuery : DELAY? = " + delay)
        if float(delay)==0:
            query_file.write("---> 3458A DELAY? Matches With Python \n")
        else:
            query_file.write("---> 3458A DELAY? Does Not Match Python \n")
        query_file.close()
        lfilter = str(self.instrument.query('LFILTER?')) #Check to see if 3458A is set to 'LFILTER' mode
        query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store LFILTER? validation in txt file.
        query_file.write("\nQuery : LFILTER? = " + lfilter)
        if lfilter[0]==str(1):
            query_file.write("---> 3458A LFILTER? Matches With Python \n")
        else:
            query_file.write("---> 3458A LFILTER? Does Not Match Python \n")
        query_file.close()
    self.instrument.write('NRDGS '+str(Num)+',TIMER')
    if self.CodeValidation.QueryValidate.GetValue()==True:
        nrdgs = str(self.instrument.query('NRDGS?')) #Check to see if 3458A is set to 'NRDGS' mode
        query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store NRDGS? validation in txt file.
        query_file.write("\nQuery : NRDGS? = " + nrdgs)
        if str(int(Num)) in nrdgs:
            query_file.write("---> 3458A NRDGS? Matches With Python="+str(Num) +"\n")
        else:
            query_file.write("---> 3458A NRDGS? Doesnt Match With Python. \n")
        query_file.close()
    memory_available = self.instrument.query('MSIZE?').split(',') #generates 2 element list, first element is memory available for measuremnts
    Storage = int( int(memory_available[0])/4 )
    #print(" ")
    #print("MACHINE SETUP COMPLETE")
    #print(" ")
    if Num>Storage:
        print("NOT ENOUGH VOLTMETER MEMORY FOR NEEDED SAMPLES")
        print("TRY A LARGER Aper_targ VALUE OR SMALLER Num")


    ######  PRELIMINARY COMPUTATIONS


    Bw_corr=FNVmeter_bw(Freq,Range)
    ##error estimate##
    if Range>120:
        Base = 15.0   #self heating and base error
    else:
        Base = 10.0
    #Vmeter_bw IS ERROR DUE TO UNCERTAINTY IN KNOWING THE HIGH FREQUENCY
    #RESPONSE OF THE DCV FUNCTION FOR VARIOUS RANGES AND FREQUENCIES
    #UNCERTAINTY IS 30% AND THIS ERROR IS RANDOM
    X1=FNVmeter_bw(Freq,Range)
    X2=FNVmeter_bw(Freq*1.3,Range)
    Vmeter_bw=int((1e+6)*abs(X2-X1))   #error due to meter band width, bw.
    #Aper_er IS THE DCV GAIN ERROR FOR VARIOUS A/D APERTURES
    #THIS ERROR IS SPECIFIED IN A GRAPH ON PAGE 11 OF THE DATA SHEET
    #THIS ERROR IS RANDOM
    Aper_er=int(1.0*.002/Aper)       #GAIN UNCERTAINTY - SMALL A/D APERTURE
    if Aper_er>30 and Aper>=1e-5:
        Aper_er=30
    if Aper<1e-5:
        Aper_er=10+int(0.0002/Aper)
    #Sincerr IS THE ERROR DUE TO THE APERTURE TIME NOT BEING PERFECTLY KNOWN
    #THIS VARIATION MEANS THAT THE Sinc CORRECTION TO THE SIGNAL FREQUENCY
    #IS NOT PERFECT.  ERROR COMPONENTS ARE CLOCK FREQ UNCERTAINTY(0.01%)
    #AND SWITCHING TIMING (50ns).  THIS ERROR IS RANDOM.
    X=np.pi*Aper*Freq       #USED TO CORRECT FOR A/D APERTURE ERROR
    Sinc = np.sin(X)/X
    Y=np.pi*Freq*(Aper*1.0001+5.0e-8)
    Sinc2=np.sin(Y)/Y
    Sincerr=int(1e+6*abs(Sinc2-Sinc))   #APERTURE UNCERTAINTY ERROR
    #Dist IS ERROR DUE TO UP TO 1% DISTORTION OF THE INPUT WAVEFORM
    #IF THE INPUT WAVEFORM HAS 1% DISTORTION, THE ASSUMPTION IS MADE
    #THAT THIS ENERGY IS IN THE THIRD HARMONIC.  THE APERTURE CORRECTION,
    #WHICH IS MADE ONLY FOR THE FUNDAMENTAL FREQUENCY WILL THEN BE
    #INCORRECT.  THIS ERROR IS RETURNED SEPERATELY.
    Sinc3=np.sin(3*X)/3/X      #SINC CORRECTION NEEDED FOR 3rd HARMONIC
    Harm_er=abs(Sinc3-Sinc)
    Dist=np.sqrt(1.0+(0.01*(1+Harm_er))**2)-np.sqrt(1.0+0.01**2)
    Dist=int(Dist*1e+6)
    ##Tim_er IS ERROR DUE TO MISTIMING.  IT CAN BE SHOWN THAT IF A
    ##BURST OF Num SAMPLES ARE USED TO COMPUTE THE RMS VALUE OF A SINEWAVE
    ##AND THE SIZE OF THIS BURST IS WITHIN 50ns*Num OF AN INTEGRAL NUMBER
    ##OF PERIODS OF THE SIGNAL BEING MEASURED, AN ERROR IS CREATED
    ##BOUNDED BY 100ns/4/Tsamp.  THIS ERROR IS DUE TO THE 100ns QUANTIZATION
    ##LIMITATION OF THE HP3458A TIME BASE.  IF THIS ERROR WERE ZERO, THEN
    ##Num*Tsamp= INTEGER/Freq, BUT WITH THIS ERROR UP TO 50ns OF TIMEBASE
    ##ERROR IS PRESENT PER SAMPLE, THEREFORE TOTAL TIME ERROR=50ns*Num
    ##THIS ERROR CAN ONLY ACCUMULATE UP TO 1/2 *Tsamp, AT WHICH POINT THE
    ##ERROR IS BOUNDED BY 1/4/Num
    ##THIS ERROR IS FURTHER REDUCED BY USING THE LEVEL TRIGGER
    ##TO SPACE Nbursts AT TIME INCREMENTS OF 1/Nbursts/Freq.  THIS
    ##REDUCTION IS SHOWN AS 20*Nbursts BUT IN FACT IS USUALLY MUCH BETTER
    ##THIS ERROR IS ADDED ABSOLUTELY TO THE Err CALCULATION
    #(Err,Dist_er,Freq,Range,Num,Aper,Nbursts) sent params
    #(Err,Dist,Freq,Range,Num,Aper,Nbursts) recieved renamed
    Tim_er=int((1e+6)*1e-7/4/(Aper+(3.0e-5))/20.0)#ERROR DUE TO HALF CYCLE ERROR
    Limit=int((1e+6)/4.0/Num/20.0)
    if Tim_er>Limit:
        Tim_er=Limit
    ##Noise IS THE MEASUREMENT TO MEASUREMENT VARIATIONS DUE TO THE
    ##INDIVIDUAL SAMPLES HAVING NOISE.  THIS NOISE IS UNCORRELATED AND
    ##IS THEREFORE REDUCED BY THE SQUARE ROOT OF THE NUMBER OF SAMPLES
    ##THERE ARE Nbursts*Num SAMPLES IN A MEASUREMENT.  THE SAMPLE NOISE IS
    ##SPECIFIED IN THE GRAPH ON PAGE 11 OF THE DATA SHEET.  THIS GRAPH
    ##SHOWS 1 SIGMA VALUES, 2 SIGMA VALUES ARE COMPUTED BELOW.
    ##THE ERROR ON PAGE 11 IS EXPRESSED AS A % OF RANGE AND IS MULTIPLIED
    ##BY 10 SO THAT IT CAN BE USED AS % RDG AT 1/10 SCALE.
    ##ERROR IS ADDED IN AN ABSOLUTE FASHION TO THE Err CALCULATION SINCE
    ##IT WILL APPEAR EVENTUALLY IF A MEASUREMENT IS REPEATED OVER AND OVER
    Noiseraw=0.9*np.sqrt(0.001/Aper)       #1 SIGMA NOISE AS PPM OF RANGE
    Noise=Noiseraw/np.sqrt(Nbursts*Num)  #REDUCTION DUE TO MANY SAMPLES
    Noise=10.0*Noise                   #NOISE AT 1/10 FULL SCALE
    Noise=2.0*Noise                    #2 SIGMA
    if Range<=0.12:               #NOISE IS GREATER ON 0.1 V RANGE
        Noise=7.0*Noise                  #DATA SHEET SAYS USE 20, BUT FOR SMALL
        Noiseraw=7.0*Noiseraw            #APERTURES, 7 IS A BETTER NUMBER
    Noise=int(Noise)+2.0                  #ERROR DUE TO SAMPLE NOISE
    ##Df_err IS THE ERROR DUE TO THE DISSIPATION FACTOR OF THE P.C. BOARD
    ##CAPACITANCE LOADING DOWN THE INPUT RESISTANCE.  THE INPUT RESISTANCE
    ##IS 10K OHM FOR THE LOW VOLTAGE RANGES AND 100K OHM FOR THE HIGH VOLTAGE
    ##RANGES (THE 10M OHM INPUT ATTENUATOR).  THIS CAPACITANCE HAS A VALUE
    ##OF ABOUT 15pF AND A D.F. OF ABOUT 1.0%.  IT IS SWAMPED BY 120pF
    ##OF LOW D.F. CAPACITANCE (POLYPROPALENE CAPACITORS) ON THE
    ##LOW VOLTAGE RANGES WHICH MAKES FOR AN EFFECTIVE D.F. OF ABOUT .11%.
    ##THIS CAPACITANCE IS SWAMPED BY 30pF LOW D.F. CAPACITANCE ON THE
    ##HIGH VOLTAGE RANGES WHICH MAKES FOR AN EFFECTIVE D.F. OF .33%.
    ##THIS ERROR IS ALWAYS IN THE NEGATIVE DIRECTION, SO IS ADDED ABSOLUTELY
    if Range<=12:
        Rsource=10000
        Cload=1.33e-10
        Df=1.1e-3          #0.11%
    else:
        Rsource=1.0e+5
        Cload=5.00e-11
        Df=3.3e-3
    Df_err=2*np.pi*Rsource*Cload*Df*Freq
    Df_err=int(1.0e+6*Df_err)#ERROR DUE TO TO PC BOARD DIELECTRIC ABSORBTION
    #Err IS TOTAL ERROR ESTIMATION.  RANDOM ERRORS ARE ADDED IN RSS FASHION
    Err=np.sqrt(Base**2+Vmeter_bw**2+Aper_er**2+Sincerr**2)
    Err=int(Err+Df_err+Tim_er+Noise)
##    print("SIGNAL FREQUENCY(Hz)= "+str(Freq))
##    print("Number of samples in each of "+str(Nbursts)+" bursts= "+str(Num))
##    print("Sample spacing(sec)= "+str(Tsamp))
##    print("A/D Aperture(sec)= "+str(Aper))
##    print("Measurement bandwidth(Hz)= "+str(int(5/Aper)/10))
##    print("ESTIMATED TOTAL SINEWAVE MEASUREMENT UNCERTAINTY(ppm)= "+str(Err))
##    print("ADDITIONAL ERROR FOR 1% DISTORTION(3rd HARMONIC)(ppm)= "+str(Dist))
##    print("NOTE: ERROR ESTIMATE ASSUMES (ACAL DCV) PERFORMED RECENTLY(24HRS)")

    ###measuremnts###
    Sum=0
    Sumsq=0
    
    
    
    for I in range(0,int(Nbursts)):
        self.instrument.write('DISP OFF,MEASURING')
        Delay=float(I)/Nbursts/Freq+(1e-6)
        self.instrument.write('DELAY '+str(Delay))
        if self.CodeValidation.QueryValidate.GetValue()==True:
            delay = str(self.instrument.query('DELAY?')) #Check to see if 3458A is set to 'DELAY' mode
            query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store DELAY? validation in txt file.
            query_file.write("\nQuery : DELAY? = " + delay)
            query_file.write("---> 3458A Compare if DELAY? With Python Delay="+str(Delay)+"\n")
            query_file.close()
        self.instrument.write('TIMER '+str(Tsamp))
        if self.CodeValidation.QueryValidate.GetValue()==True:
            timer = float(self.instrument.query('TIMER?')) #Check to see if 3458A is set to 'TIMER' mode
            query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store TIMER? validation in txt file.
            query_file.write("\nQuery : TIMER? = " + str(timer) + " \n")
            query_file.write("---> 3458A Compare if TIMER? matches with Python="+str(Tsamp)+"\n")
            query_file.close()
    
        #make measuremnts
        self.instrument.write('MEM FIFO;MFORMAT DINT') #first in first out for memeory 
        mem = str(self.instrument.query('MEM?')) #Check to see if 3458A is set to 'MEM' mode
        if self.CodeValidation.QueryValidate.GetValue()==True:
            query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store MEM? validation in txt file.
            query_file.write("\nQuery : MEM? = " + mem)
            if mem[0]==str(2):
                query_file.write("---> 3458A MEM? matches with Python \n")
            else:
                query_file.write("---> 3458A MEM? does not match with Python \n")
            query_file.close()
            mformat = str(self.instrument.query('MFORMAT?')) #Check to see if 3458A is set to 'MFORMAT' mode
            query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store MFORMAT? validation in txt file.
            query_file.write("\nQuery : MFORMAT? = " + mformat)
            if mformat[0]==str(3):
                query_file.write("---> 3458A MFORMAT? matches with Python \n")
            else:
                query_file.write("---> 3458A MFORMAT? does not match with Python \n")
            query_file.close()
        
        #clears memeory, sets to 4 bytes per reading
        self.instrument.write('TARM SGL')
        if self.CodeValidation.QueryValidate.GetValue()==True:
            tarm1 = str(self.instrument.query('TARM?')) #Check to see if 3458A is set to 'TARM' mode
            query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store TARM? validation in txt file.
            query_file.write("\nQuery : TARM? = " + tarm1)
            if tarm1[0]==str(4):
                query_file.write("---> 3458A TARM? matches with Python. SGL Becomes HOLD. \n")
            else:
                query_file.write("---> 3458A TARM? does not match with Python. SGL Doesnt Become HOLD. \n")
            query_file.close()
        self.instrument.write('MMATH STAT')#Just makes the 3458A machine calculate the SDEV,MEAN,NSAMP,UPPER,LOWER. So no need to validate this.
        
        Sdev = float(self.instrument.query('RMATH SDEV'))
        Mean = float(self.instrument.query('RMATH MEAN'))
        
        single_burst = []
        
        if MeanV==True:
            self.instrument.write('OFORMAT ASCII')
            if self.CodeValidation.QueryValidate.GetValue()==True:
                oformat = str(self.instrument.query('OFORMAT?')) #Check to see if 3458A is set to 'OFORMAT' mode
                query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Store OFORMAT? validation in txt file.
                query_file.write("\nQuery : OFORMAT? = " + oformat)
                if oformat[0]==str(1):
                    query_file.write("---> 3458A OFORMAT? matches with Python. \n")
                else:
                    query_file.write("---> 3458A OFORMAT? matches with Python. \n")
                query_file.close()
            self.instrument.write('RMEM 1,'+str(Num)+',1')
            #print(self.instrument.query('RMEM 1,'+str(Num)+',1'))
            #single_burst = self.instrument.read_raw()
            #single_burst = self.instrument.query_ascii_values('RMEM 1,'+str(Num)+',1')
            #print repr(single_burst)
            #print repr(self.instrument.query('RMEM 1,'+str(Num)+',1'))
            
            self.instrument.write('DISP OFF,READING_MEM')
            for l in range(1,Num+1):
                reading = self.instrument.query_ascii_values('RMEM '+str(l)+',1,1')
                #print repr(reading)
                #reading = self.instrument.read()
                #print(reading)
                single_burst.append(reading) #removes non-float characters from return string
                #print(single_burst[l-1])
            self.MemArray.append(single_burst)


        Sdev=Sdev*np.sqrt((Num-1.0)/Num)     #CORRECT SDEV FORMULA
        
        Sumsq=Sumsq+Sdev*Sdev+Mean*Mean
        Sum=Sum+Mean
        Temp=Sdev*Bw_corr/Sinc
        Temp=Range/(1e+7)*int(Temp*(1e+7)/Range)#6 DIGIT TRUNCATION
        
        #print("Sdev: "+str(Sdev))
        #print("Mean: "+str(Mean))
        #print("RMS: "+str(np.sqrt(Sdev**2+Mean**2)))
        #print(Temp)
        I=I+1
    Dcrms=np.sqrt(Sumsq/Nbursts)
    #print("RMS value: "+str(Dcrms))
    #print(" ")
    Dc=Sum/Nbursts
    Acrms=np.sqrt(Dcrms**2-Dc**2)
    Acrms=Acrms*Bw_corr/Sinc  #CORRECT A/D Aper AND Vmeter B.W.
    Dcrms=np.sqrt(Acrms*Acrms+Dc*Dc)
    Acrms=Range/1e+7*int(Acrms*1e+7/Range+0.5)    #6 DIGIT TRUNCATION
    Dcrms=Range/1e+7*int(Dcrms*1e+7/Range+0.5)    #6 DIGIT TRUNCATION
    self.instrument.write('DISP ON')

    if Acdc == True:
        self.AcdcArray.append(Dcrms)
        print("AcDcrms = " + str(Dcrms))
    if Ac == True:
        self.AcArray.append(Acrms)
        print("Acrms = " + str(Acrms))
    if MeanV == True:
        rms_initial = np.sqrt(np.mean(MemArray**2))
        ratio = Dcrms/rms_initial #ratio of true amplitude to measured amplitude
        MemArray = MemArray*ratio #corrected array of all data
        self.MeanArray.append(np.mean(np.abs(self.MemArray))) #save into array of means
        print("MeanV = " + str(np.mean(np.abs(self.MemArray))))
    #if Mem == True:
        #print(self.MemArray)
        #SaveMem("Mem = " + str(self.MemArray))
    print("Sinc = " + str(Sinc))
    print("Bw_corr = " + str(Bw_corr))
    print("Freq = " + str(Freq))
    print("#Samples = " + str(Num))
    print("Tsamp = " + str(Tsamp))
    if Force==False:
        print("Ncycles = " + str(Ncycle))
    print("Nbursts = " + str(Nbursts)+"\n")
    if self.CodeValidation.validate==False:
        #The following Code records data measurements to an Excel Sheet.
        wb=load_workbook(filename) #Open Excel file
        ws=wb.active #Make it active to work in
    
    if Acdc==True:
        if self.CodeValidation.validate==False:
            ws.cell(row=row_inc,column=1,value=Dcrms) #First column
    if Ac == True:
        if self.CodeValidation.validate==False:
            ws.cell(row=row_inc,column=1,value=Acrms) #First column
    if MeanV == True:
        if self.CodeValidation.validate==False:
            ws.cell(row=row_inc,column=1,value=np.mean(np.abs(self.MemArray))) #First column
    if self.CodeValidation.validate==False:
        ws.cell(row=row_inc,column=2,value=Bw_corr) #Second column
        ws.cell(row=row_inc,column=3,value=Freq) #Third column
        ws.cell(row=row_inc,column=4,value=Num) #Fourth column
        ws.cell(row=row_inc,column=5,value=Tsamp) #Fifth column
        if Force==False:
            ws.cell(row=row_inc,column=6,value=Ncycle) #Sixth column
        ws.cell(row=row_inc,column=7,value=Nbursts) #Eight column
        ws.cell(row=row_inc,column=8,value=self.ForceParameters.GetValue()) #Seventh column
        date_and_time = str(time.asctime())
        newtime = date_and_time[11:20]
        ws.cell(row=row_inc,column=9,value=newtime) #Ninth column
        row_inc+=1 #Go to next row
        wb.template=False #Make sure Excel file is saved as document not template
        wb.save(filename) #Save file with same name
    
    self.times.append(float(time.time()))

