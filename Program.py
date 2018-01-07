#importing wx files
import wx
 
import gui
import SwerleinFreq as alg
#import numpy as np
import matplotlib
matplotlib.use('WXAgg')
#from matplotlib.figure import Figure
#from matplotlib.backends.backend_wxagg import \
#        FigureCanvasWxAgg as FigCanvas, \
#        NavigationToolbar2WxAgg as NavigationToolbar
from openpyxl import Workbook
from openpyxl import load_workbook
#import time
import datetime
import winsound
import visa

#Inherit from the CodeValidation Frame in gui.py.
class CodeValid(gui.CodeValidation):
    #Constructor
    def __init__(self,parent):
        gui.CodeValidation.__init__(self,parent)
        self.validate = False #Validate option not active (by default).
    
    def CodeValidationOnClose( self, event ):
        #Reset Validate Options to Default value (False).
        self.validate = False 
        self.QueryValidate.SetValue(False)
        self.MathValidate.SetValue(False)
        self.Show(False) #Stop displaying code validation frame.

    def RunValidation(self):
        remote = visa.ResourceManager() #Gets remote access of the 5730A Calibrator 
        supply = remote.open_resource('GPIB0::16::INSTR')
        
        Acdc = frame.ACDC.GetValue()
        Ac = frame.AC.GetValue()
        Mean = frame.MEAN.GetValue()
##         Mem = self.memory.GetValue()
        
        alg.OnStart(frame,frame.Port.GetValue())
        n = int(frame.NumReadings.GetValue())
        frame.date=str(datetime.date.today())
        voltage_list=[0.65]
        for v in voltage_list:
            if self.QueryValidate.GetValue()==True:
                query_file = open("Code_Validation (Ver 2.0)"+str(frame.date)+".txt","a") #Create code_validation in txt file.
                query_file.write("\n ---------------------"+str(v)+"V "+str(frame.date)+"---------------------\n")
                query_file.close()
            if self.MathValidate.GetValue()==True:
                filename2="Math_Validation (Ver 2.0)"+str(frame.date)+".xlsx"
                wb = Workbook()
                wb.template=False
                wb.save(filename2)
                frame.filename2=filename2
                wb=load_workbook(frame.filename2) #Open Excel file
                ws=wb.active #Make it active to work in
                #Add tile labels to theie respective locations
                ws.cell(row=1,column=1,value="Python Calculated Variable Values")
                ws.cell(row=2,column=1,value="RMS"),ws.cell(row=3,column=1,value="Range"),ws.cell(row=4,column=1,value="Expect_Freq"),ws.cell(row=5,column=1,value="Freq"),ws.cell(row=6,column=1,value="Aper"),ws.cell(row=7,column=1,value="Tsamp"),ws.cell(row=8,column=1,value="Submeas_time")
                ws.cell(row=9,column=1,value="Burst_time"),ws.cell(row=10,column=1,value="Approxnum"),ws.cell(row=11,column=1,value="Ncycle"),ws.cell(row=12,column=1,value="Nharm"),ws.cell(row=13,column=1,value="Tsamptemp"),ws.cell(row=14,column=1,value="Num"),ws.cell(row=15,column=1,value="K")
                ws.cell(row=16,column=1,value="Bw_corr"),ws.cell(row=17,column=1,value="Base"),ws.cell(row=18,column=1,value="X1"),ws.cell(row=19,column=1,value="X2"),ws.cell(row=20,column=1,value="Vmeter_bw"),ws.cell(row=21,column=1,value="Aper_er"),ws.cell(row=22,column=1,value="X")
                ws.cell(row=23,column=1,value="Sinc"),ws.cell(row=24,column=1,value="Y"),ws.cell(row=25,column=1,value="Sinc2"),ws.cell(row=26,column=1,value="Sincerr"),ws.cell(row=27,column=1,value="Sinc3"),ws.cell(row=28,column=1,value="Harm_er"),ws.cell(row=29,column=1,value="Dist")
                ws.cell(row=30,column=1,value="Tim_er"),ws.cell(row=31,column=1,value="Limit"),ws.cell(row=32,column=1,value="Noiseraw"),ws.cell(row=33,column=1,value="Noise"),ws.cell(row=34,column=1,value="Rsource"),ws.cell(row=35,column=1,value="Cload"),ws.cell(row=36,column=1,value="Df")
                ws.cell(row=37,column=1,value="Df_err"),ws.cell(row=38,column=1,value="Err"),ws.cell(row=39,column=1,value="Sum"),ws.cell(row=40,column=1,value="Sumsq"),ws.cell(row=41,column=1,value="Delay"),ws.cell(row=42,column=1,value="Sdev"),ws.cell(row=43,column=1,value="Mean")
                ws.cell(row=44,column=1,value="Temp"),ws.cell(row=45,column=1,value="Dcrms"),ws.cell(row=46,column=1,value="Dc"),ws.cell(row=47,column=1,value="Acrms")
                wb.template=False #Make sure Excel file is saved as document not template
                wb.save(frame.filename2) #Save file with same name
                supply.write("OUT "+str(v)+"V,50.0HZ")
            for i in range (0,n):
                alg.run(frame,float(frame.Harmonics.GetValue()),float(frame.Bursts.GetValue()),Acdc,Ac,Mean,frame.row_inc,"" )
                i = i+1
                frame.row_inc = frame.row_inc + 1
        
        winsound.Beep(2500,1000)
        frame.row_inc=2 #Reset to Row 2 for next Excel file.
        print("\n DONE-----DONE-----DONE-----DONE-----DONE")

    def ValidateOnButtonClick(self, event):
        self.validate = True #Make validate option active.
        if self.QueryValidate.GetValue()==True or self.MathValidate.GetValue()==True: 
            self.RunValidation()
        
    

#Inherit from the Swerlein Frame created in gui.py.
class CalcFrame(gui.Swerlein):
    
    #constructor
    def __init__(self,parent):
        #initialize parent class
        self.row_inc=2 #Always start at 2. The 1 row has the column titles
        try:
            alg.OnStart(self,self.Port.GetValue())
        except Exception:
            pass
        gui.Swerlein.__init__(self,parent)
        self.CodeValidation = CodeValid(None) #Create new code validation frame. All options set to False by default.

    def RunFunc(self,event):
        #Create a new workbook in Excel.
        if self.CodeValidation.validate==False:
            self.date=str(datetime.date.today())
            filename='data'+str(self.date)+str(self.Excel_Filename.GetValue())+'.xlsx'
            #Comment out following three lines if wanting to continue working in an existing Excel file.
            #Remember to change the filename parameter initialized previously to the existing filename
            #As well as changing self.row_inc to the last+1 elemt in the existing file
            wb = Workbook()
            wb.template=False
            wb.save(filename)
            self.filename=filename
        #remote = visa.ResourceManager() #Gets remote access of the 5730A Calibrator 
        #supply = remote.open_resource('GPIB0::16::INSTR')
        
        Acdc = self.ACDC.GetValue()
        Ac = self.AC.GetValue()
        Mean = self.MEAN.GetValue()
##         Mem = self.memory.GetValue()
        
        alg.OnStart(self,self.Port.GetValue())
        n = int(self.NumReadings.GetValue())
        if self.CodeValidation.validate==False:
            wb=load_workbook(self.filename) #Open Excel file
            ws=wb.active #Make it active to work in
            #Add tiles to the columns
            ws.cell(row=1,column=1,value="Voltage")
            ws.cell(row=1,column=2,value="Bw_corr")
            ws.cell(row=1,column=3,value="Frequency")
            ws.cell(row=1,column=4,value="# of Samples")
            ws.cell(row=1,column=5,value="Sample Spacing")
            ws.cell(row=1,column=6,value="# of 1/Freq to Sample")
            ws.cell(row=1,column=7,value="# of bursts")
            ws.cell(row=1,column=8,value="Parameters Forced?")
            ws.cell(row=1,column=9,value="Time")
            wb.template=False #Make sure Excel file is saved as document not template
            wb.save(self.filename) #Save file with same name
        
        #voltage_list=[0.65,0.512,0.001,1.0,0.023]
        #for v in voltage_list:
        if self.CodeValidation.QueryValidate.GetValue()==True:
            query_file = open("Code_Validation (Ver 2.0)"+str(self.date)+".txt","a") #Create code_validation in txt file.
            query_file.write("\n ---------------------"+str(v)+"V "+str(self.date)+"---------------------\n")
            query_file.close()
        #supply.write("OUT "+str(v)+"V,50.0HZ")
        for i in range (0,n):
            alg.run(self,float(self.Harmonics.GetValue()),float(self.Bursts.GetValue()),Acdc,Ac,Mean,self.row_inc,self.filename )
            i = i+1
            self.row_inc = self.row_inc + 1
        
        winsound.Beep(2500,1000)
        self.row_inc=2 #Reset to Row 2 for next Excel file.
        print("\n DONE-----DONE-----DONE-----DONE-----DONE")

    def StopFunc(self,event):
        alg.reset(self)
         
    def AboutOnMenuSelection( self, event ):
        gui.About(None).Show(True) #Display a window containing information about the software.
        
    def QueryValidationOnMenuSelection( self, event ):
        self.CodeValidation = CodeValid(None)
        self.CodeValidation.Show(True) #Display the already created CodeValidation frame.

#mandatory in wx, create an app, False stands for not deteriction stdin/stdout
#refer manual for details
app = wx.App(False)

#create an object of CalcFrame
frame = CalcFrame(None)
#show the frame
frame.Show(True)
#start the applications
app.MainLoop()
